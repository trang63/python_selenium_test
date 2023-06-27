import openpyxl
import os


class HomePageData:
    # Get the current script file's directory
    current_directory = os.path.dirname(os.path.abspath(__file__))

    # Navigate upwards in the directory hierarchy until the project root is reached
    while not os.path.isfile(os.path.join(current_directory, "README.md")):
        current_directory = os.path.dirname(current_directory)

    project_root = current_directory

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(HomePageData.project_root + "/TestData/pythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

    @staticmethod
    def getAllTestData():
        ListData = []
        book = openpyxl.load_workbook(HomePageData.project_root + "/TestData/pythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if "testcase" in sheet.cell(row=i, column=1).value.lower():
                Dict = {}
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                ListData.append(Dict)
        return ListData


# print(HomePageData.getTestData("Testcase2"))
print(HomePageData.getAllTestData())
