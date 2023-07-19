import openpyxl
import os


class DataUtilities:
    @staticmethod
    def getTestData(test_case_name, file_path):
        Dict = {}
        book = openpyxl.load_workbook(file_path)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

    @staticmethod
    def getAllTestData(file_path):
        ListData = []
        book = openpyxl.load_workbook(file_path)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if "testcase" in sheet.cell(row=i, column=1).value.lower():
                Dict = {}
                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                ListData.append(Dict)
        return ListData
